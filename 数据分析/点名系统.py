# C:\Bin\environment\python
# -*- coding: utf-8 -*-
# @Time : 2022/9/25 9:49
# @Author : south(南风)
# @File : 点名系统.py
# Describe: 
# -*- coding: utf-8 -*-
import random
import time

import openpyxl
import tkinter as tk


def get_students_name():
    # 学生名单中需要有"姓名"列
    workbook = openpyxl.load_workbook('grade.xlsx')
    table = workbook.active
    # print(table)
    rows, cols = table.max_row, table.max_column
    name_col = 0
    for col in range(cols):
        if table.cell(1, col + 1).value == 'sname':
            name_col = col
            break
    students_name = [table.cell(row + 1, name_col + 1).value for row in range(1, rows)
                     if table.cell(row + 1, name_col + 1).value is not None]
    return students_name


# a = get_students_name()
# print(a)
def call_lucky_student(var):
    """点名"""
    global is_run
    if is_run:
        return
    is_run = True
    start = time.time()
    choice_student(var, start)


def choice_student(var, start):
    global is_run
    global name1
    show_member = random.choice(name1)
    name = show_member[0]
    for zi in show_member[1:]:
        name += ' ' + zi
    var.set(name)
    end = time.time()
    if is_run and end - start <= 5:
        window.after(30, choice_student, var, start)
    else:
        is_run = False
        return


if __name__ == '__main__':
    is_run = False
    name1 = get_students_name()
    window = tk.Tk()
    window.geometry('400x600')
    window.title('\t 学 生 点 名 系 统')
    # 添加背景图片
    bg_png = tk.PhotoImage(file="../img/background.png")
    bg_label = tk.Label(window, image=bg_png)
    bg_label.pack()
#     # 添加显示框
    var = tk.StringVar(value='公平 公正 公开')
    show_label1 = tk.Label(window, textvariable=var, justify='left', anchor=tk.CENTER, width=16,
                           height=2, font='楷体 -40 bold', foreground='white', bg='#1C86EE')
    show_label1.place(anchor=tk.NW, x=30, y=150)
#     # 添加点名按钮
    button_png = tk.PhotoImage(file='../img/resume_nor.png')
    button = tk.Button(window, text='点 名', compound='center', font='楷体 -30 bold',
                       foreground='#9400D3', image=button_png,
                       command=lambda: call_lucky_student(var))
    button.place(anchor=tk.NW, x=130, y=300)

#     # 显示窗口
    window.mainloop()
    if is_run is False:
        button_png = tk.PhotoImage(file='../img/resume_nor.png')
    else:
        button_png = tk.PhotoImage(file='../img/pause_nor.png')
#
