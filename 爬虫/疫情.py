# C:\Bin\environment\python
# -*- coding: utf-8 -*-
# @Time : 2022/10/25 19:11
# @Author : south(南风)
# @File : 疫情.py
# Describe: 
# -*- coding: utf-8 -*-
import requests
from lxml import etree
import openpyxl
#
# wb = openpyxl.load_workbook('疫情.xlsx')
# sheet3 = wb.create_sheet(title='Sheet3', index=0)

# wb = openpyxl.Workbook()
# # 利用openpyxl.Workbook()函数创建新的workbook（工作簿）对象，就是创建新的空的Excel文件
# sheet = wb.active
# # wb.active就是获取这个工作簿的活动表，通常就是第一个工作表
# sheet.title = '疫情'
# # 可以用.title给工作表重命名。现在第一个工作表的名称就会由原来默认的“sheet1”改为"电影工作表"
# sheet['A1'] = '高风险'
# 往A1的单元格中写入了'熊出没'




# url = 'http://m.bj.bendibao.com/news/gelizhengce/fengxianmingdan.php?src=baidu'
# headers = {
#     'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.3 Mobile/15E148 Safari/604.1'
# }
#
# response = requests.get(url=url, headers=headers)
# with open('12.html', 'w', encoding='utf-8') as f:
#     f.write(response.text)

with open('2.html', 'r', encoding='utf-8') as f:
    content = f.read()

xml = etree.HTML(content)
xian1 = xml.xpath('/html/body/div[1]/div[2]/div[6]/div')
# print(len(xian1))
a = 0
for i in range(1, len(xian1) + 1, 2):
    # city = xml.xpath(f'/html/body/div[1]/div[2]/div[6]/div/div[{i}]/div')
    # for j in range(1, len(city) + 1):
    xian = xml.xpath(f'/html/body/div[1]/div[2]/div[6]/div[{i}]/div/p//text()')
    num = xml.xpath(f'/html/body/div[1]/div[2]/div[6]/div[{i}]/div/span//text()')
    print(num[0])
    a = a + int(num[0].strip("个"))
        # xian = xml.xpath(f'/html/body/div[1]/div[2]/div[6]/div/div[{i}]/div[{j}]/div[1]/p/span//text()')
        # if len(xian) == 3:
        #     addr = xian[0] + xian[1]
        # else:
        #     addr = xian[0]
    addr = xian[0].strip(" \n") + xian[1].strip(" \n") + xian[2].strip(" \n")
    jvti = xml.xpath(f'/html/body/div[1]/div[2]/div[6]/div[{i + 1}]/ul/li/span//text()')
    for z in jvti:
        if z.strip(" ") is not None:
            list1 = []
            print(addr + z)
            list1.append((addr + z).strip(" "))
            # sheet3.append(list1)
            list1.pop()

print(a)
# /html/body/div[1]/div[2]/div[6]/div/div[1]/div[2]
# wb.save('疫情.xlsx')

